from collections.abc import Iterable
from typing import NoReturn, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

import glytrait
from glytrait.config import Config
from glytrait.exception import InputError, StructureParseError
from glytrait.glycan import NGlycan
from glytrait.trait import TraitFormula


def read_input(file: str) -> tuple[list[NGlycan] | None, pd.DataFrame]:
    """Read the input file.

    The first column should be "Composition", with condensed glycan composition notations.
    The second column should be "Structure", with glycan structure strings like glycoCT.
    The rest columns should be the abundance of the glycans in different samples.
    The "Structure" column is optional.

    Args:
        file (str): The input csv or xlsx file.

    Returns:
        tuple[list[NGlycan], pd.DataFrame]: The glycans and the abundance table. The glycans are
            represented by NGlycan objects. The abundance table has the glycans as columns and
            samples as rows. If "Structure" column is not provided, the glycans will be None.

    Raises:
        InputError: If the input file format is wrong.
    """
    df = pd.read_csv(file)
    df = df.dropna(how="all")
    has_structure = "Structure" in df.columns
    _check_input(df, has_structure)
    df = df.set_index("Composition")

    if has_structure:
        glycans = []
        for glycan_id, structure in zip(df.index, df["Structure"]):
            try:
                glycan = NGlycan.from_glycoct(structure)
            except StructureParseError as e:
                raise StructureParseError(glycan_id + ": " + str(e))
            else:
                glycans.append(glycan)
        abundance_table = df.drop(columns=["Structure"]).T
    else:
        glycans = None
        abundance_table = df.T

    return glycans, abundance_table


def read_group(file: str) -> pd.Series:
    """Read the group file.

    The first column should be the sample IDs and the second column should be the group names.

    Args:
        file (str): The group file.

    Returns:
        groups (pd.Series): The group names, with sample IDs as index.
    """
    df = pd.read_csv(file, index_col=0)
    if df.shape[1] != 1:
        raise InputError("The group file should only have two columns.")
    groups = df.squeeze()
    return groups


def read_structure(file: str, compositions: Iterable[str]) -> list[NGlycan]:
    """Read the structure file.

    The first column should be the glycan compositions, and the second column should be the
    glycan structures.
    Only the structures of the glycans in `compositions` will be read, and the order of the
    structures will be the same as that of `compositions`.

    Args:
        file (str): The structure file.
        compositions (Iterable[str]): The compositions of the glycans to be read.

    Returns:
        list[NGlycan]: The glycans.

    Raises:
        InputError: If the input file format is wrong, or some compositions are not found in the
            structure file.
    """
    df = pd.read_csv(file, index_col=0)
    if df.shape[1] != 1:
        raise InputError("The structure file should only have two columns.")
    structures = df.squeeze()
    glycans = []
    for composition in compositions:
        try:
            glycan = NGlycan.from_glycoct(structures[composition])
        except StructureParseError as e:
            raise StructureParseError(composition + ": " + str(e))
        except KeyError:
            raise InputError(composition + " is not found in the structure file.")
        else:
            glycans.append(glycan)
    return glycans


def _check_input(df: pd.DataFrame, has_structure: bool) -> NoReturn:
    """Check the input dataframe."""
    if df.columns[0] != "Composition":
        raise InputError("The first column of the input file should be Composition.")
    if df["Composition"].duplicated().sum() > 0:
        raise InputError("There are duplicated Compositions in the input file.")

    if has_structure:
        if df.columns[1] != "Structure":
            raise InputError("The second column of the input file should be Structure.")
        if df["Structure"].duplicated().sum() > 0:
            raise InputError("There are duplicated Structures in the input file.")

    if np.any(df.iloc[:, 2 if has_structure else 1:].dtypes != "float64"):
        raise InputError("The abundance columns in the input file should be numeric.")

    if df.isna().sum().sum() > 0:
        raise InputError("There are missing values in the input file.")


def write_output(
    config: Config,
    derived_traits: pd.DataFrame,
    direct_traits: pd.DataFrame,
    meta_prop_df: pd.DataFrame,
    formulas: list[TraitFormula],
    groups: Optional[pd.Series] = None,
    hypothesis_test: Optional[pd.DataFrame] = None,
) -> None:
    """Write the output file.

    Args:
        config (Config): The configuration.
        derived_traits (pd.DataFrame): The trait values, with samples as index and trait names as
            columns.
        direct_traits (pd.DataFrame): The normalized abundance table, with samples as index and
            glycan IDs as columns.
        meta_prop_df (pd.DataFrame): The table of meta properties generated by
            `build_meta_property_table`.
        formulas (list[TraitFormula]): The trait formulas.
        groups (pd.Series): The group names, with sample IDs as index. Optional.
        hypothesis_test (Optional[pd.DataFrame], optional): The hypothesis test results. Optional.
    """
    wb = Workbook()
    ws0: Worksheet = wb.active
    ws0.title = "Summary"
    _write_summary(ws0, config, direct_traits, derived_traits, groups)

    # The trait table
    ws1 = wb.create_sheet("Trait values")
    _write_trait_values(ws1, direct_traits, derived_traits)

    # The trait definitions
    ws2 = wb.create_sheet("Trait definitions")
    _write_trait_defination(ws2, formulas)

    # The meta properties
    ws3 = wb.create_sheet("Meta properties")
    _write_meta_properties(ws3, meta_prop_df)

    # The hypothesis test results
    if hypothesis_test is not None:
        ws4 = wb.create_sheet("Hypothesis test results")
        _write_hypothesis_test(ws4, hypothesis_test)

    wb.save(config.get("output_file"))


def _write_summary(
    ws: Worksheet,
    config: Config,
    direct_triats: pd.DataFrame,
    derived_traits: pd.DataFrame,
    groups: Optional[pd.Series] = None,
) -> None:
    """Write the summary sheet."""
    ws["A1"] = "GlyTrait version"
    ws["B1"] = glytrait.__version__

    ws["A2"] = "Options"
    ws.merge_cells("A2:B2")
    ws["A2"].font = ws["A2"].font.copy(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"] = "Input file"
    ws["B3"] = config.get("input_file")
    ws["A4"] = "Output file"
    ws["B4"] = config.get("output_file")
    ws["A5"] = "Sialic acid linkage"
    ws["B5"] = config.get("sia_linkage")
    ws["A6"] = "Imputation method"
    ws["B6"] = config.get("impute_method")
    ws["A7"] = "Glycan filter ratio"
    ws["B7"] = config.get("filter_glycan_max_na")
    ws["A8"] = "Invalid trait filtering"
    ws["B8"] = config.get("filter_invalid_traits")
    ws["A9"] = "Group file"
    ws["B9"] = config.get("group_file")
    ws["A10"] = "Formula file"
    ws["B10"] = config.get("formula_file")
    ws["A11"] = "Structure file"
    ws["B11"] = config.get("structure_file")
    ws["A12"] = "Database"
    ws["B12"] = config.get("database")
    ws["A13"] = "Input file has structure"
    ws["B13"] = config.get("_has_struc_col")

    ws["A14"] = "Result Overview"
    ws.merge_cells("A14:B14")
    ws["A14"].font = ws["A14"].font.copy(bold=True)
    ws["A14"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A15"] = "Num. of derict traits"
    ws["B15"] = len(direct_triats.columns)
    ws["A16"] = "Num. of derived traits"
    ws["B16"] = len(derived_traits.columns)
    ws["A17"] = "Total num. of traits"
    ws["B17"] = len(direct_triats.columns) + len(derived_traits.columns)

    if groups is not None:
        ws["A18"] = "Statistical Analysis"
        ws.merge_cells("A15:B15")
        ws["A18"].font = ws["A15"].font.copy(bold=True)
        ws["A18"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A19"] = "Num. of groups"
        ws["B19"] = len(groups.unique())
        ws["A30"] = "Hypothesis testing method"
        ws["B20"] = (
            "Mann-Whitney U Test"
            if len(groups.unique()) == 2
            else "Kruskal-Wallis H Test"
        )

    ws.column_dimensions["A"].width = 27
    for cell in ws["B"]:
        cell.alignment = Alignment(horizontal="left")


def _write_trait_values(
    ws: Worksheet,
    direct_traits: pd.DataFrame,
    derived_traits: pd.DataFrame,
) -> None:
    """Write the trait values sheet."""
    combined_df = pd.concat([direct_traits, derived_traits], axis=1)

    for row in dataframe_to_rows(combined_df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    for cell in ws[1][1:]:
        cell.font = cell.font.copy(bold=True)
        cell.border = cell.border.copy(bottom=Side(border_style="double"))
    ws.insert_rows(1)
    ws.merge_cells(
        start_row=1,
        start_column=2,
        end_row=1,
        end_column=len(direct_traits.columns) + 1,
    )
    ws.cell(1, 2).value = "Direct traits"
    ws.merge_cells(
        start_row=1,
        start_column=len(direct_traits.columns) + 2,
        end_row=1,
        end_column=len(combined_df.columns) + 1,
    )
    ws.cell(1, len(direct_traits.columns) + 2).value = "Derived traits"
    for cell in ws[1][1:]:
        cell.font = cell.font.copy(bold=True)
        cell.border = cell.border.copy(bottom=Side(border_style="thin"))


def _write_trait_defination(ws: Worksheet, formulas: list[TraitFormula]) -> None:
    """Write the trait defination sheet."""
    ws.append(["Trait Name", "Description"])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for formula in formulas:
        ws.append([formula.name, formula.description])


def _write_meta_properties(ws: Worksheet, meta_prop_df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(meta_prop_df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)


def _write_hypothesis_test(ws: Worksheet, hypothesis_test: pd.DataFrame) -> None:
    for row in dataframe_to_rows(hypothesis_test, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
